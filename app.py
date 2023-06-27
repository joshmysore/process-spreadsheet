import os
import pandas as pd
import numpy as np
from flask import Flask, render_template, redirect, url_for, send_from_directory, abort, flash, request
from werkzeug.utils import secure_filename
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, NamedStyle, Side, Border
import math as math

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['PROCESSED_FOLDER'] = 'processed/'
app.secret_key = "super secret key"

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xls', 'xlsx', 'xlsm'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Ningún archivo seleccionado')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('Ningún archivo seleccionado')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            processed_files = process_spreadsheet(file_path)
            return redirect(url_for('download_success', filename=processed_files))
        else:
            flash('Los tipos de archivo permitidos son xls, xlsx, xlsm')
            return redirect(request.url)

    return render_template ('index.html')

@app.route('/downloads/<filename>')
def download_file(filename):
    return send_from_directory(app.config['PROCESSED_FOLDER'], filename, as_attachment=True)

@app.route('/download')
def download_success():
    return render_template('download.html')

def process_spreadsheet(file_path):
    # Create a new workbook to store all typologies
    wb = Workbook()

    # Define style for odd rows
    odd_row_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    odd_row_style = NamedStyle(name="odd_row_style", fill=odd_row_fill)

    # Define style for even rows
    even_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    even_row_style = NamedStyle(name="even_row_style", fill=even_row_fill)

    # Add these styles to the workbook
    wb.add_named_style(odd_row_style)
    wb.add_named_style(even_row_style)

    # Read all sheets from the current excel file
    dfs = pd.read_excel(file_path, sheet_name=None)

    # Process each sheet
    for sheet_name, df in dfs.items():
        # Check if the apartment has "Sí" in column "Estudio" and then set "Tipología" to "Estudio". Then, if the apartment does not have "Sí" in column "Estudio", set "Tipología" to the concatenation of "Habitaciones" and "Baños".
        df['Tipología'] = np.where(df['Estudio'] == 'Si', 'Estudio',
                                    df['Habitaciones'].astype(str) + 'D' + df['Baños'].astype(str) + 'B')

        # Remove duplicates based on latitude
        df.drop_duplicates(subset='Latitud', keep='first', inplace=True)

        # Create 'm2 totales' column
        df['m2 totales'] = df[['Superficie', 'Mts Total', 'Mts Útil', 'Mts Total Imp', 'Mts Útil Imp']].max(axis=1)

        # Create empty 'rangos' column
        df['Rangos'] = ''

        # Delete columns 'Url Busconido' and 'Descripción' and 'F. Desactivación'
        df.drop(columns=['Url Busconido', 'Descripción', 'F. Desactivación'], inplace=True)

        # Find the position of 'Precio ($)' column
        price_index = df.columns.get_loc('Precio ($)')

        # Reorder the dataframe columns
        df = df.reindex(df.columns.tolist()[:price_index + 1] + ['m2 totales', 'Rangos'] + df.columns.tolist()[price_index + 1:-2], axis=1)

        # Group by Tipología
        grouped = df.groupby('Tipología')

        # For each typology, calculate statistics and add to the same sheet
        for typology, group in grouped:
            # Calculate min and max values of m2 totales
            min_m2 = min(group['m2 totales'])
            max_m2 = max(group['m2 totales'])

            # Calculate the number of filter ranges. 
            num_ranges = math.ceil((max_m2 - min_m2) / 10)
            filter_ranges = [(min_m2 + i * 10, min_m2 + (i + 1) * 10) for i in range(num_ranges)]

            # Using filter_ranges, add ranges to the rangos columns that correspond to the m2 totales column.
            group['Rangos'] = pd.cut(group['m2 totales'], 
                bins=[range[0] for range in filter_ranges]+[max_m2+1], 
                labels=[f'{range[0]}-{range[1]}' for range in filter_ranges], 
                include_lowest=True)

            # Don't forget to assign the modified group back to the df dataframe
            df.loc[group.index, :] = group

            # Create a new sheet with the typology as the name
            sheet = wb.create_sheet(typology)

            # Write the grouped dataframe to the sheet
            for r in dataframe_to_rows(group, index=False, header=True):
                sheet.append(r)

            # Apply styles to rows
            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=len(group) + 1), 2):
                for cell in row:
                    if i % 2 == 0:
                        cell.style = odd_row_style
                    else:
                        cell.style = even_row_style
                    # Add borders between columns and rows. Add color #d3d3d3 as hex.
                    cell.border = Border(left=Side(border_style='thin', color='d3d3d3'),
                                        right=Side(border_style='thin', color='d3d3d3'),
                                        top=Side(border_style='thin', color='d3d3d3'),
                                        bottom=Side(border_style='thin', color='d3d3d3'))

            # Calculate row number for the stats box
            row_num_stats = len(group) + 4

            # Define stats box formatting
            bold_font = Font(bold=True)
            blue_fill = PatternFill(start_color='9ab7e6', end_color='9ab7e6', fill_type='solid')

            # List of labels
            labels = ['Promedio:', 'Moda:',
                        'Rango Mínimo:', 'Rango Máximo:', 'Percentil 80:', 'Percentil 85:', 'Percentil 90:', 'Percentil 95:']

            # Write the column and row name in the stats box
            for i in range(len(labels) + 1):
                if i == 0:  # For the first row
                    sheet.cell(row=row_num_stats, column=3, value='N. Precio ($)').font = bold_font
                    sheet.cell(row=row_num_stats, column=4, value='N. m2 totales').font = bold_font
                    sheet.cell(row=row_num_stats, column=3).fill = blue_fill
                    sheet.cell(row=row_num_stats, column=4).fill = blue_fill
                else:  # For subsequent rows
                    sheet.cell(row=row_num_stats + i, column=2, value=labels[i - 1]).font = bold_font
                    sheet.cell(row=row_num_stats + i, column=2).fill = blue_fill

            # Find the columns for "Precio ($)" and "m2 totales"
            precio_column = None
            m2_column = None
            for col, cell in enumerate(sheet[1], start=1):
                if cell.value == "Precio ($)":
                    precio_column = col
                elif cell.value == "m2 totales":
                    m2_column = col

            if precio_column is None or m2_column is None:
                print(f"Columns not found for 'Precio ($)' and 'm2 totales'")
                continue

            # Define abstract functions and their arguments
            # The key is the function number for the AGGREGATE function in Excel.
            # The value is the arguments for the functions.
            functions = {
                1: [None],  # AVERAGE
                13: [None], # MODE.SNGL
                5: [None],  # MIN
                4: [None],  # MAX
                18: ['0.8', '0.85', '0.9', '0.95']  # PERCENTILE
            }

            # Loop over the functions and apply them
            i = 0
            for function, args in functions.items():
                for arg in args:
                    # Handle functions without arguments differently
                    if arg is None:
                        formula_precio = f'=AGGREGATE({function}, 5, {sheet.cell(row=2, column=precio_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=precio_column).column_letter}{row_num_stats - 1})'
                        formula_m2 = f'=AGGREGATE({function}, 5, {sheet.cell(row=2, column=m2_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=m2_column).column_letter}{row_num_stats - 1})'
                    else:
                        formula_precio = f'=AGGREGATE({function}, 5, {sheet.cell(row=2, column=precio_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=precio_column).column_letter}{row_num_stats - 1}, {arg})'
                        formula_m2 = f'=AGGREGATE({function}, 5, {sheet.cell(row=2, column=m2_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=m2_column).column_letter}{row_num_stats - 1}, {arg})'

                    sheet.cell(row=row_num_stats + 1 + i, column=3, value=formula_precio).number_format = '$#,##0.00'
                    sheet.cell(row=row_num_stats + 1 + i, column=4, value=formula_m2).number_format = '#,##0.00'
                    i += 1

            # Adjust column widths based on text length for all columns
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                column_letter = column_cells[0].column_letter
                sheet.column_dimensions[column_letter].width = max_length + 4

                # Adjust width for the "ID" column (Column A)
                max_length_id = max(len(str(cell.value)) for cell in sheet['A'])
                sheet.column_dimensions['A'].width = max_length_id + 4

            # Bold and fill in gray column headers in the top row
            for cell in sheet[1]:
                cell.font = bold_font
                cell.fill = blue_fill

            # Find the 'Rangos' column
            rangos_column = None
            for col, cell in enumerate(sheet[1], start=1):
                if cell.value == 'Rangos':
                    rangos_column = col
                    break

            if rangos_column is None:
                print(f"Column not found for 'Rangos'")
            else:
                # The filter should cover all rows in the 'Rangos' column
                # We define the filter range from the first row (header) up to the last row in the sheet
                # We use the column_letter method to get the column letter from the column number
                rangos_column_letter = sheet.cell(row=1, column=rangos_column).column_letter
                sheet.auto_filter.ref = f'{rangos_column_letter}1:{rangos_column_letter}{sheet.max_row}'

    # Remove default sheet created by Workbook
    del wb['Sheet']
    
    # Update the process_spreadsheet function to save processed files in PROCESSED_FOLDER
    processed_file_name = 'processed_' + os.path.basename(file_path)
    processed_file_path = os.path.join(app.config['PROCESSED_FOLDER'], processed_file_name)
    # Save processed workbook to processed_file_path instead of just processed_file_name
    # wb.save(processed_file_name)
    wb.save(processed_file_path)
    return processed_file_name

if __name__ == '__main__':
    app.run(debug=True)

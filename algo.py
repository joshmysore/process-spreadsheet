import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, NamedStyle, Side, Border
import math as math
import logging
from datetime import datetime
import warnings

# Define función para seleccionar el archivo
def setup_process():
    # Ask user for filename once and store it in a variable
    input('Presione Enter para seleccionar el archivo de Excel a procesar...')
    root = tk.Tk()
    root.withdraw()
    selected_file = filedialog.askopenfilename()
    selected_file_name = os.path.basename(selected_file)
    # Get current date
    now = datetime.now()
    # Format date as string
    current_date = now.strftime("%Y_%d_%m")
    folder_name = 'resultados_{0}_{1}'.format(selected_file_name, current_date)

    # make folder
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        logging.info('Carpeta creada en {0}'.format(os.getcwd()))

    # config para logging
    logging.basicConfig(
        level=logging.INFO, 
        format='%(asctime)s - %(levelname)s - %(message)s', 
        handlers=[
            # name log after the file selected as such debug_{filename of selected file}_time.log and put it in folder_name
            logging.FileHandler("{0}/debug_{1}_{2}.log".format(folder_name, selected_file_name, current_date)),
            logging.StreamHandler()
        ]
    )
    # statement saying log file has been created and saved whereever it is found
    logging.info('Archivo de log creado en {0}'.format(os.getcwd()))

    # change the working directory to the new folder
    os.chdir(folder_name)

    logging.info('Empezando el programa')

    return selected_file

# Define función para todo
def process_spreadsheet(selected_file):
    # Logging
    logging.info(f'Archivo seleccionado: {selected_file}')

    # Crea un nuevo archivo de Excel
    wb = Workbook()

    # Define estilo para las celdas de cabecera
    odd_row_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    odd_row_style = NamedStyle(name="odd_row_style", fill=odd_row_fill)
    even_row_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    even_row_style = NamedStyle(name="even_row_style", fill=even_row_fill)

    # Añade el estilo a los estilos del archivo
    wb.add_named_style(odd_row_style)
    wb.add_named_style(even_row_style)

    logging.info(f'Procesando el archivo {selected_file}')
    # Compueba que el archivo seleccionado es un archivo de Excel
    if not selected_file.endswith((".xls", ".xlsx")):
        logging.info(f'"{selected_file}" no es un archivo de Excel.')

    try:
        # Comprueba que el archivo no está abierto en otro programa
        with open(selected_file, "r") as f:
            pass
    except IOError:
        logging.info(f'"{selected_file}" está abierto en otro programa.')


    try:
        # Comprueba que el archivo no está vacío
        dfs = pd.read_excel(selected_file, sheet_name=None)
        first_sheet = list(dfs.keys())[0]
        if dfs[first_sheet].empty:
            logging.info(f'"{selected_file}" está vacío. Saltando este archivo.')
    except Exception as e:
        logging.info(
            f'No se puede leer "{selected_file}". Error: {str(e)}.'
        )

    # logging statement that states checks that have been passed
    logging.info(f'Archivo {selected_file} ha pasado los checks de ser un archivo de Excel con contenido.')

    # Comprueba que el archivo contiene todas las columnas necesarias
    required_columns = [
        "Estudio",
        "Habitaciones",
        "Baños",
        "Latitud",
        "Superficie",
        "Mts Total",
        "Mts Útil",
        "Mts Total Imp",
        "Mts Útil Imp",
        "Url Busconido",
        "Descripción",
        "F. Desactivación",
        "Precio ($)",
    ]

    # Lee el archivo de Excel en un DataFrame
    dfs = pd.read_excel(selected_file, sheet_name=None)

    for sheet_name, df in dfs.items():
        if not set(required_columns).issubset(df.columns):
            print(
                f'Hoja "{sheet_name}" en "{selected_file}" no contiene las columnas requeridas. Saltando este archivo.'
            )
            dfs.pop(sheet_name)

    for sheet_name, df in dfs.items():
        if not set(required_columns).issubset(df.columns):
            print(
                f'Hoja "{sheet_name}" en "{selected_file}" no contiene las columnas requeridas. Saltando este archivo.'
            )
            dfs.pop(sheet_name)
            continue
        else: 
            logging.info(f'Hoja "{sheet_name}" en "{selected_file}" contiene las columnas requeridas.')
        # Compueba que las columnas "Habitaciones" y "Baños" contienen valores numéricos
        if df["Habitaciones"].dtype not in ["int64", "float64"] or df[
            "Baños"
        ].dtype not in ["int64", "float64"]:
            print(
                f'Hoja "{sheet_name}" en "{selected_file}" contiene los datos incorrectos para  "Habitaciones" y/o "Baños". Saltando este archivo.'
            )
            dfs.pop(sheet_name)
            continue
        df["Habitaciones"] = (df["Habitaciones"] // 1).fillna("NaN")
        df["Baños"] = (df["Baños"] // 1).fillna("NaN")
        if df["Habitaciones"].min() < 1 or df["Baños"].min() < 1:
            print(
                f'Hoja "{sheet_name}" en "{selected_file}" contiene los datos incorrectos para  "Habitaciones" y/o "Baños". Saltando este archivo.'
            )
            dfs.pop(sheet_name)
            continue

    # Procesa cada hoja del archivo de Excel
    for sheet_name, df in dfs.items():
        # Crea tipología
        df["Tipología"] = np.where(
            df["Estudio"] == "Si",
            "Estudio",
            df["Habitaciones"].astype(str) + "D" + df["Baños"].astype(str) + "B",
        )
        logging.info(f'Columna Tipología creada con: {df["Tipología"]}')
        # Borra duplicados
        df.drop_duplicates(subset="Latitud", keep="first", inplace=True)
        logging.info(f'Duplicados borrados basado en la columna Latitud')

        # Calcula m2 totales y rangos
        df["m2 totales"] = calc_m2_totales(df)
        df["Rangos"] = ""
        logging.info(f'Columna m2 totales creada con: {df["m2 totales"]}')
        logging.info(f'Columna Rangos creada con: {df["Rangos"]}')

        # Borra columnas innecesarias
        df.drop(
            columns=["Url Busconido", "Descripción", "F. Desactivación"],
            inplace=True,
        )
        logging.info(f'Columnas innecesarias borradas: Url Busconido, Descripción, F. Desactivación')

        # Calcula el índice de la columna Precio ($)
        price_index = df.columns.get_loc("Precio ($)")

        # Reordena las columnas
        df = df.reindex(
            df.columns.tolist()[: price_index + 1]
            + ["m2 totales", "Rangos"]
            + df.columns.tolist()[price_index + 1 : -2],
            axis=1,
        )
        logging.info(f'Columnas reordenadas')

        # Agrupa por tipología
        grouped = df.groupby("Tipología")
        logging.info(f'Columnas agrupadas por Tipología: {grouped}')

        # Por cada tipología, calcula estadísticas
        for typology, group in grouped:
            # Convierte la columna m2 totales a numérico
            group["m2 totales"] = pd.to_numeric(group["m2 totales"], errors='coerce')

            # Elimina filas con m2 totales vacíos
            group = group.dropna(subset=["m2 totales"])

            # Calcula el min y max de m2 totales
            min_m2 = min(group["m2 totales"])
            max_m2 = max(group["m2 totales"])
            logging.info(f'Calculando min y max de m2 totales: {min_m2} y {max_m2}')

            # Si el min y max son iguales, solo hay un rango
            if min_m2 == max_m2:
                num_ranges = 1
            else:
                # Calcula el número de rangos 
                num_ranges = math.ceil((max_m2 - min_m2) / 10)
                logging.info(f'Calculando número de rangos: {num_ranges}')

            # Crea una lista de tuplas con los rangos
            filter_ranges = [
                (min_m2 + i * 10, min_m2 + (i + 1) * 10) for i in range(num_ranges)
            ]
            logging.info(f'Creando lista de tuplas con los rangos: {filter_ranges}')

            # Añade el rango a la columna Rangos
            rangos_values = pd.cut(
                group["m2 totales"],
                bins=[range[0] for range in filter_ranges] + [max_m2 + 1],
                labels=[f"{range[0]}-{range[1]}" for range in filter_ranges],
                include_lowest=True,
            )
            logging.info(f'Añadiendo el rango a la columna Rangos.')

            # Mete el grupo en el dataframe
            df.loc[group.index, "Rangos"] = rangos_values

            # Crea nuevas hojas para cada tipología
            sheet_name = typology
            sheet_counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{typology}_{sheet_counter}"
                sheet_counter += 1
            sheet = wb.create_sheet(str(sheet_name))
            logging.info(f'Creando hoja {sheet_name}')

            # Crea estilos para las filas pares e impares
            for r in dataframe_to_rows(group, index=False, header=True):
                sheet.append(r)

            # Define estilos para las filas pares e impares
            for i, row in enumerate(
                sheet.iter_rows(min_row=2, max_row=len(group) + 1), 2
            ):
                for cell in row:
                    if i % 2 == 0:
                        cell.style = odd_row_style
                    else:
                        cell.style = even_row_style
                    # Define el borde de las celdas
                    cell.border = Border(
                        left=Side(border_style="thin", color="d3d3d3"),
                        right=Side(border_style="thin", color="d3d3d3"),
                        top=Side(border_style="thin", color="d3d3d3"),
                        bottom=Side(border_style="thin", color="d3d3d3"),
                    )
            logging.info(f'Estilos definidos para las filas pares e impares')
            # Añade los estadísticos
            calc_stats(sheet, group)
            logging.info(f'Añadiendo estadísticos')

            # Ajusta el ancho de las columnas
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                column_letter = column_cells[0].column_letter
                sheet.column_dimensions[column_letter].width = max_length + 4

                # Ajusta el ancho de la columna A
                max_length_id = max(len(str(cell.value)) for cell in sheet["A"])
                sheet.column_dimensions["A"].width = max_length_id + 4
            logging.info(f'Ajustando el ancho de las columnas')

            # Define el estilo de las celdas de estadísticas
            bold_font = Font(bold=True)
            blue_fill = PatternFill(
                start_color="9ab7e6", end_color="9ab7e6", fill_type="solid"
            )
            logging.info(f'Estilo de celdas de estadísticas definido como bold_font y blue_fill')
            # Da el diseño a la tabla
            for cell in sheet[1]:
                cell.font = bold_font
                cell.fill = blue_fill

            # Busca la columna de Rangos
            rangos_column = None
            for col, cell in enumerate(sheet[1], start=1):
                if cell.value == "Rangos":
                    rangos_column = col
                    break

            if rangos_column is None:
                print(f"Column not found for 'Rangos'")
            else:
                # Filtra la columna de Rangos
                rangos_column_letter = sheet.cell(
                    row=1, column=rangos_column
                ).column_letter
                sheet.auto_filter.ref = (
                    f"{rangos_column_letter}1:{rangos_column_letter}{sheet.max_row}"
                )

    # Elimina la hoja por defecto
    del wb["Sheet"]

    # Guarda el archivo y dice la dirección
    processed_file_name = "procesado_" + os.path.basename(selected_file)
    wb.save(processed_file_name)
    #make this print statemnet a logging one
    logging.info(f'Procesamiento de datos finalizado. Resultados guardados como" {processed_file_name}" en {os.getcwd()}')

# Define m2 totales función
def calc_m2_totales(df):
    columns = ["Superficie",
               "Mts Total",
               "Mts Útil",
               "Mts Total Imp",
               "Mts Útil Imp"]
    df["m2 totales"] = df[columns].max(axis=1)
    logging.info(f'Columnas {columns} procesadas para calcular m2 totales')
    return df["m2 totales"]

# Define función para calcular estadísticas
def calc_stats(sheet, group):
    # Define el número de filas de estadísticas
    row_num_stats = len(group) + 4
    logging.info(f'Calculando estadísticas para {len(group)} filas')

    # Define el estilo de las celdas de estadísticas
    bold_font = Font(bold=True)
    blue_fill = PatternFill(
        start_color="9ab7e6", end_color="9ab7e6", fill_type="solid"
    )
    logging.info(f'Estilo de celdas de estadísticas definido como bold_font y blue_fill')

    # Define las etiquetas de las estadísticas
    labels = [
        "Promedio:",
        "Moda:",
        "Mediana:",
        "Rango Mínimo:",
        "Rango Máximo:",
        "Percentil 80:",
        "Percentil 85:",
        "Percentil 90:",
        "Percentil 95:",
    ]
    logging.info(f'Etiquetas de estadísticas definidas: {labels}')

    # Define las fórmulas de las estadísticas
    for i in range(len(labels) + 1):
        if i == 0:  # For the first row
            sheet.cell(
                row=row_num_stats, column=3, value="N. Precio ($)"
            ).font = bold_font
            sheet.cell(
                row=row_num_stats, column=4, value="N. m2 totales"
            ).font = bold_font
            sheet.cell(row=row_num_stats, column=3).fill = blue_fill
            sheet.cell(row=row_num_stats, column=4).fill = blue_fill
        else:  # Para el resto de filas
            sheet.cell(
                row=row_num_stats + i, column=2, value=labels[i - 1]
            ).font = bold_font
            sheet.cell(row=row_num_stats + i, column=2).fill = blue_fill

    # Busca las columnas de Precio ($) y m2 totales
    precio_column = None
    m2_column = None
    for col, cell in enumerate(sheet[1], start=1):
        if cell.value == "Precio ($)":
            precio_column = col
        elif cell.value == "m2 totales":
            m2_column = col

    if precio_column is None or m2_column is None:
        print(f"Columns not found for 'Precio ($)' and 'm2 totales'")
        return

    # Define las funciones de estadísticas
    functions = {
        1: [None],  # PROMEDIO
        13: [None],  # MODA
        12: [None],  # MEDIANA
        5: [None],  # MIN
        4: [None],  # MAX
        18: ["0.8", "0.85", "0.9", "0.95"],  # PERCENTIL
    }

    # Bucle para calcular las estadísticas
    i = 0
    for function, args in functions.items():
        for arg in args:
            # Hazlo con argumentos y sin argumentos
            if arg is None:
                formula_precio = f"=AGREGAR({function}, 5, {sheet.cell(row=2, column=precio_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=precio_column).column_letter}{row_num_stats - 1})"
                formula_m2 = f"=AGREGAR({function}, 5, {sheet.cell(row=2, column=m2_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=m2_column).column_letter}{row_num_stats - 1})"
            else:
                formula_precio = f"=AGREGAR({function}, 5, {sheet.cell(row=2, column=precio_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=precio_column).column_letter}{row_num_stats - 1}, {arg})"
                formula_m2 = f"=AGREGAR({function}, 5, {sheet.cell(row=2, column=m2_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=m2_column).column_letter}{row_num_stats - 1}, {arg})"

            sheet.cell(
                row=row_num_stats + 1 + i, column=3, value=formula_precio
            ).number_format = "$#,##0.00"
            sheet.cell(
                row=row_num_stats + 1 + i, column=4, value=formula_m2
            ).number_format = "#,##0.00"
            i += 1

if __name__ == "__main__":
    selected_file = setup_process()
    process_spreadsheet(selected_file)
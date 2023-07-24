import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, NamedStyle, Side, Border
import math as math
import logging
from datetime import datetime

# Define función para seleccionar el archivo
def setup_process():
    # Configurar logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler()  # comenzar solo con el controlador de flujo
        ]
    )

    # Solicitar al usuario el nombre de archivo y guardarlo en una variable
    input('Presione Enter para seleccionar el archivo de Excel a procesar...')
    root = tk.Tk()
    root.withdraw()
    selected_file = filedialog.askopenfilename()
    selected_file_name = os.path.basename(selected_file)
    # Obtener la fecha actual
    now = datetime.now()
    # Formatear la fecha como una cadena
    current_date = now.strftime("%Y_%d_%m_%H_%M_%S")
    folder_name = 'resultados_{0}_{1}'.format(selected_file_name, current_date)

    # crear carpeta
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        logging.info('Carpeta creada en {0}'.format(os.getcwd()))

    # Ahora que existe el directorio, crear el archivo de registro en él
    file_handler = logging.FileHandler("{0}/debug_{1}_{2}.log".format(folder_name, selected_file_name, current_date))
    logging.getLogger().addHandler(file_handler)

    # declaración que indica que se ha creado y guardado el archivo de registro donde se encuentre
    logging.info('Archivo de log creado en {0}'.format(os.getcwd()))

    # cambiar el directorio de trabajo a la nueva carpeta
    os.chdir(folder_name)

    logging.info('Empezando el programa')

    return selected_file

# Define función para calcular m2 totales
def calc_m2_totales(df):
    columns = ["Superficie",
               "Mts Total",
               "Mts Útil",
               "Mts Total Imp",
               "Mts Útil Imp"]
    # Aegurarse de que solo haya números y eliminar los NAN en las columnas
    for column in columns:
        df[column] = pd.to_numeric(df[column], errors='coerce')
        # eliminar los NAN
        df[column] = df[column].fillna(0)
    # Calcular el máximo de las columnas
    df["m2 totales"] = df[columns].max(axis=1)
    logging.info(f'Columnas {columns} procesadas para calcular m2 totales')
    return df["m2 totales"]

# Define función para calcular estadísticas
def calc_stats(sheet, group):
    # Define el número de filas de estadísticas
    row_num_stats = len(group) + 4
    logging.info(f'Calculando estadísticas para {len(group)} filas')

    # Define el estilo de las celdas de estadísticas
    bold_font = Font(bold=True)
    blue_fill = PatternFill(
        start_color="9ab7e6", end_color="9ab7e6", fill_type="solid"
    )
    logging.info(f'Estilo de celdas de estadísticas definido como bold_font y blue_fill')

    # Define las etiquetas de las estadísticas
    labels = [
        "Promedio:",
        "Moda:",
        "Mediana:",
        "Rango Mínimo:",
        "Rango Máximo:",
        "Percentil 80:",
        "Percentil 85:",
        "Percentil 90:",
        "Percentil 95:",
    ]
    logging.info(f'Etiquetas de estadísticas definidas: {labels}')

    # Define las fórmulas de las estadísticas
    for i in range(len(labels) + 1):
        if i == 0:  # Para la primera fila
            sheet.cell(
                row=row_num_stats, column=3, value="N. Precio ($)"
            ).font = bold_font
            sheet.cell(
                row=row_num_stats, column=4, value="N. m2 totales"
            ).font = bold_font
            sheet.cell(row=row_num_stats, column=3).fill = blue_fill
            sheet.cell(row=row_num_stats, column=4).fill = blue_fill
        else:  # Para el resto de filas
            sheet.cell(
                row=row_num_stats + i, column=2, value=labels[i - 1]
            ).font = bold_font
            sheet.cell(row=row_num_stats + i, column=2).fill = blue_fill

    # Buscar las columnas de Precio ($) y m2 totales
    precio_column = None
    m2_column = None
    for col, cell in enumerate(sheet[1], start=1):
        if cell.value == "Precio ($)":
            precio_column = col
        elif cell.value == "m2 totales":
            m2_column = col

    if precio_column is None or m2_column is None:
        print(f"Columnas no encontradas para 'Precio ($)' y 'm2 totales'")
        return

    # Definir las funciones de estadísticas
    functions = {
        1: [None],  # PROMEDIO
        13: [None],  # MODA
        12: [None],  # MEDIANA
        5: [None],  # MIN
        4: [None],  # MAX
        18: ["0.8", "0.85", "0.9", "0.95"],  # PERCENTIL
    }

    # Bucle para calcular las estadísticas
    i = 0
    for function, args in functions.items():
        for arg in args:
            # Hacerlo con argumentos y sin argumentos
            if arg is None:
                formula_precio = f"=AGREGAR({function}, 5, {sheet.cell(row=2, column=precio_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=precio_column).column_letter}{row_num_stats - 1})"
                formula_m2 = f"=AGREGAR({function}, 5, {sheet.cell(row=2, column=m2_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=m2_column).column_letter}{row_num_stats - 1})"
            else:
                formula_precio = f"=AGREGAR({function}, 5, {sheet.cell(row=2, column=precio_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=precio_column).column_letter}{row_num_stats - 1}, {arg})"
                formula_m2 = f"=AGREGAR({function}, 5, {sheet.cell(row=2, column=m2_column).column_letter}2:{sheet.cell(row=row_num_stats - 1, column=m2_column).column_letter}{row_num_stats - 1}, {arg})"

            sheet.cell(
                row=row_num_stats + 1 + i, column=3, value=formula_precio
            ).number_format = "$#,##0.00"
            sheet.cell(
                row=row_num_stats + 1 + i, column=4, value=formula_m2
            ).number_format = "#,##0.00"
            i += 1

# Define función para comprobar la validez del archivo
def check_file_validity(selected_file):
    # Comprobar si selected_file es un archivo de Excel
    if not selected_file.endswith((".xls", ".xlsx")):
        logging.info(f'"{selected_file}" no es un archivo de Excel.')
        return False

    try:
        # Comprobar si el archivo no está abierto por otro programa
        with open(selected_file, "r") as f:
            pass
    except IOError:
        logging.info(f'"{selected_file}" está abierto en otro programa.')
        return False

    try:
        # Comprobar si el archivo no está vacío
        dfs = pd.read_excel(selected_file, sheet_name=None)
        first_sheet = list(dfs.keys())[0]
        if dfs[first_sheet].empty:
            logging.info(f'"{selected_file}" está vacío. Se omitirá este archivo.')
            return False
    except Exception as e:
        logging.info(f'No se puede leer "{selected_file}". Error: {str(e)}.')
        return False

    # Declaración de registro que indica las comprobaciones que se han pasado
    logging.info(f'El archivo {selected_file} ha pasado las comprobaciones de ser un archivo de Excel con contenido.')
    return True

# Define función para leer y preprocesar el archivo
def read_and_preprocess_file(selected_file):
    # Leer el archivo de Excel en un DataFrame
    dfs = pd.read_excel(selected_file, sheet_name=None)

    # Columnas requeridas
    required_columns = [
        "Estudio",
        "Habitaciones",
        "Baños",
        "Latitud",
        "Superficie",
        "Mts Total",
        "Mts Útil",
        "Mts Total Imp",
        "Mts Útil Imp",
        "Url Busconido",
        "Descripción",
        "F. Desactivación",
        "Precio ($)",
    ]

    # Comprobar si cada hoja en el archivo contiene todas las columnas requeridas
    for sheet_name, df in dfs.items():
        if not set(required_columns).issubset(df.columns):
            logging.info(
                f'La hoja "{sheet_name}" en "{selected_file}" no contiene las columnas requeridas. Se omitirá esta hoja.'
            )
            dfs.pop(sheet_name)

    # Comprobar si las columnas "Habitaciones" y "Baños" contienen valores numéricos
    for sheet_name, df in dfs.items():
        if df["Habitaciones"].dtype not in ["int64", "float64"] or df[
            "Baños"
        ].dtype not in ["int64", "float64"]:
            logging.info(
                f'La hoja "{sheet_name}" en "{selected_file}" contiene datos incorrectos para "Habitaciones" y/o "Baños". Se omitirá esta hoja.'
            )
            dfs.pop(sheet_name)
            continue

        # Convertir "Habitaciones" y "Baños" a enteros, reemplazar NaN con "NaN"
        df["Habitaciones"] = (df["Habitaciones"] // 1).fillna("NaN")
        df["Baños"] = (df["Baños"] // 1).fillna("NaN")

        # Si "Habitaciones" y "Baños" contienen valores inferiores a 1, registrar un mensaje y omitir la hoja
        if df["Habitaciones"].min() < 1 or df["Baños"].min() < 1:
            logging.info(
                f'La hoja "{sheet_name}" en "{selected_file}" contiene datos incorrectos para "Habitaciones" y/o "Baños". Se omitirá esta hoja.'
            )
            dfs.pop(sheet_name)
            continue

    return dfs

# Define función para crear y aplicar estilos
def create_and_apply_styles(wb):
    # Crear estilos para el libro de trabajo
    odd_row_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    odd_row_style = NamedStyle(name="odd_row_style", fill=odd_row_fill)
    even_row_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    even_row_style = NamedStyle(name="even_row_style", fill=even_row_fill)

    # Agregar los estilos a los estilos con nombre del libro de trabajo
    wb.add_named_style(odd_row_style)
    wb.add_named_style(even_row_style)

    return odd_row_style, even_row_style

# Define función para procesar las hojas
def process_sheet(dfs, wb, odd_row_style, even_row_style):
    for sheet_name, df in dfs.items():   
        # Crear tipología
        df["Tipología"] = np.where(
            df["Estudio"] == "Si",
            "Estudio",
            df["Habitaciones"].astype(str) + "D" + df["Baños"].astype(str) + "B",
        )
        logging.info(f'Columna Tipología creada con: {df["Tipología"]}')
        
        # Eliminar duplicados
        df.drop_duplicates(subset="Latitud", keep="first", inplace=True)
        logging.info(f'Duplicados borrados basados en la columna Latitud')

        # Calcular m2 totales y rangos
        df["m2 totales"] = calc_m2_totales(df)
        df["Rangos"] = ""
        logging.info(f'Columna m2 totales creada con: {df["m2 totales"]}')
        logging.info(f'Columna Rangos creada con: {df["Rangos"]}')

        # Eliminar columnas innecesarias
        df.drop(
            columns=["Url Busconido", "Descripción", "F. Desactivación"],
            inplace=True,
        )
        logging.info(f'Columnas innecesarias borradas: Url Busconido, Descripción, F. Desactivación')

        # Calcular el índice de la columna Precio ($)
        price_index = df.columns.get_loc("Precio ($)")

        # Reorganizar las columnas
        df = df.reindex(
            df.columns.tolist()[: price_index + 1]
            + ["m2 totales", "Rangos"]
            + df.columns.tolist()[price_index + 1 : -2],
            axis=1,
        )
        logging.info(f'Columnas reorganizadas')

        # Agrupar por tipología
        grouped = df.groupby("Tipología")
        logging.info(f'Columnas agrupadas por Tipología: {grouped}')

        # Para cada tipología, calcular estadísticas
        for typology, group in grouped:
            # Calcular el mínimo y máximo de m2 totales
            min_m2 = min(group["m2 totales"])
            max_m2 = max(group["m2 totales"])
            logging.info(f'Calculando mínimo y máximo de m2 totales: {min_m2} y {max_m2}')

            # Si el mínimo y máximo son iguales, solo hay un rango
            if min_m2 == max_m2:
                num_ranges = 1
            else:
                # Calcular el número de rangos 
                num_ranges = math.ceil((max_m2 - min_m2) / 10)
                logging.info(f'Calculando número de rangos: {num_ranges}')

            # Crear una lista de tuplas con los rangos
            filter_ranges = [
                (min_m2 + i * 10, min_m2 + (i + 1) * 10) for i in range(num_ranges)
            ]
            logging.info(f'Creando lista de tuplas con los rangos: {filter_ranges}')

            # Agregar el rango a la columna Rangos
            group["Rangos"] = pd.cut(
                group["m2 totales"],
                bins=[range[0] for range in filter_ranges] + [max_m2 + 1],
                labels=[f"{range[0]}-{range[1]}" for range in filter_ranges],
                include_lowest=True,
            )
            logging.info(f'Añadiendo el rango a la columna Rangos: {group["Rangos"]}')

            # Colocar el grupo en el DataFrame
            df.loc[group.index, :] = group

            # Crear nuevas hojas para cada tipología
            sheet_name = typology
            sheet_counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{typology}_{sheet_counter}"
                sheet_counter += 1
            sheet = wb.create_sheet(str(sheet_name))
            logging.info(f'Creando hoja {sheet_name}')

            # Crear estilos para las filas pares e impares
            for r in dataframe_to_rows(group, index=False, header=True):
                sheet.append(r)

            # Establecer estilos para las filas pares e impares
            for i, row in enumerate(
                sheet.iter_rows(min_row=2, max_row=len(group) + 1), 2
            ):
                for cell in row:
                    if i % 2 == 0:
                        cell.style = odd_row_style
                    else:
                        cell.style = even_row_style
                    # Establecer el borde de las celdas
                    cell.border = Border(
                        left=Side(border_style="thin", color="d3d3d3"),
                        right=Side(border_style="thin", color="d3d3d3"),
                        top=Side(border_style="thin", color="d3d3d3"),
                        bottom=Side(border_style="thin", color="d3d3d3"),
                    )
            logging.info(f'Estilos definidos para las filas pares e impares')

            # Agregar las estadísticas
            calc_stats(sheet, group)
            logging.info(f'Añadiendo estadísticos')

            # Ajustar el ancho de las columnas
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                column_letter = column_cells[0].column_letter
                sheet.column_dimensions[column_letter].width = max_length + 4

                # Ajustar el ancho de la columna A
                max_length_id = max(len(str(cell.value)) for cell in sheet["A"])
                sheet.column_dimensions["A"].width = max_length_id + 4
            logging.info(f'Ajustando el ancho de las columnas')

            # Establecer el estilo de las celdas de estadísticas
            bold_font = Font(bold=True)
            blue_fill = PatternFill(
                start_color="9ab7e6", end_color="9ab7e6", fill_type="solid"
            )
            logging.info(f'Estilo de celdas de estadísticas definido como bold_font y blue_fill')
            # Dar diseño a la tabla
            for cell in sheet[1]:
                cell.font = bold_font
                cell.fill = blue_fill

            # Buscar la columna Rangos
            rangos_column = None
            for col, cell in enumerate(sheet[1], start=1):
                if cell.value == "Rangos":
                    rangos_column = col
                    break

            if rangos_column is None:
                print(f"Columna no encontrada para 'Rangos'")
            else:
                # Filtrar la columna Rangos
                rangos_column_letter = sheet.cell(
                    row=1, column=rangos_column
                ).column_letter
                sheet.auto_filter.ref = (
                    f"{rangos_column_letter}1:{rangos_column_letter}{sheet.max_row}"
                )

    return wb

# Define función para guardar el archivo
def save_workbook(wb, selected_file):
    # Guardar el libro de trabajo e imprimir la dirección
    processed_file_name = "procesado_" + os.path.basename(selected_file)
    wb.save(processed_file_name)
    logging.info(f'Procesamiento de datos finalizado. Resultados guardados como "{processed_file_name}" en {os.getcwd()}')
    return processed_file_name